import os
import re
from openpyxl import load_workbook, Workbook

output_dir = "Output"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

wb = load_workbook('Input/Input.xlsx')
sheet = wb.active

headers = [cell.value for cell in sheet[1]]
output_file_header = ["Name", "group", "Hours", "Cost", "OT hours", "OT Cost", "Total Cost", "notes"]

job_code_index = headers.index('Job Code')
employee_code_index = headers.index('Employee')
regular_code_index = headers.index('Regular Hours')
overtime_code_index = headers.index('Overtime Hours')

all_files = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[job_code_index]:
        file_path = "{}/{}.xlsx".format(output_dir,row[job_code_index])
        overtime_value = row[overtime_code_index]
        overtime_pay = float(overtime_value) * 49.5 if overtime_value else ""
        cost_ = float(row[regular_code_index]) * 33
        total_cost = (cost_ + overtime_pay) if overtime_pay else cost_
        
        
        # employ_name=str(row[employee_code_index]).split(",") if str(row[employee_code_index]) else str(row[employee_code_index])
        data=row[employee_code_index].split(",")
        
        full_name=""
        print(data) 
       
        if len(data)>=2:
            full_name="{},{}".format(data[1],data[0])
        else:
            full_name=row[employee_code_index]
            
            
            
        if os.path.exists(file_path):
            existing_wb = load_workbook(file_path)
            existing_sheet = existing_wb.active
            existing_sheet.append([full_name, "", row[regular_code_index], cost_, row[overtime_code_index], overtime_pay, total_cost])
            existing_wb.save(file_path)
        else:
            new_wb = Workbook()
            ws = new_wb.active
            ws.append(output_file_header)
            ws.append([full_name, "", row[regular_code_index], cost_, row[overtime_code_index], overtime_pay, total_cost])
            new_wb.save(file_path)
            file_info = {
                "file_path": file_path,
                "file_name": row[job_code_index],
            }
            all_files.append(file_info)

for file in all_files:
    wb = load_workbook(file["file_path"])
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    hours_index = headers.index('Hours')
    overtime_hours_index = headers.index("OT hours")
    total_cost_index = headers.index("Total Cost")
    
    total_hours = 0
    total_overtime_hours = 0
    total_cost = 0
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        overtime_hours = float(row[overtime_hours_index]) if row[overtime_hours_index] else 0
        total_hours += float(row[hours_index])
        total_overtime_hours += overtime_hours
        total_cost_value = float(row[total_cost_index]) if row[total_cost_index] else 0
        total_cost += total_cost_value
    
    sheet.append(["", "", "", ""])
    sheet.append(["", "", "", ""])
    sheet.append(["", "", total_hours, "", total_overtime_hours * 1.5])
    sheet.append(["", "", "", ""])
    sheet.append(["", "", "", ""])

    pattern = r'\b\d{8}\b'
    match = re.search(pattern, file['file_name'])
    if match:
        eight_digit_number = match.group()
        name = file['file_name'].replace(eight_digit_number, "") if eight_digit_number else file['file_name']
        sheet.append(["", "", "Job Name:", name])
        sheet.append(["", "", "P.O.:", eight_digit_number])
    else:
        sheet.append(["", "", "Job Name:", file['file_name'], ""])
        sheet.append(["", "", "P.O.:", ""])
    
    sheet.append(["", "", "", ""])
    sheet.append(["", "", "", ""])
    sheet.append(["", "", "", "", total_hours + (total_overtime_hours * 1.5), "", total_cost])
    
    wb.save(file["file_path"])
